import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1,title='Second Sheet')

ws['A1'] = 'Invoice'

ws['A1'].font = Font(name= 'Times New Roman', size= 24, bold=True)

fontobject = Font(name= 'Times New Roman', size=24, bold=True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')


ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'
ws['A8'].font = fontobject

ws['B8'] = '=SUM(B2:B7)'

ws.column_dimensions['A'].width = 15

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb["ProduceReport"]

write_sheet["A1"] = 'Produce'
write_sheet["B1"] = 'Cost per pound'
write_sheet["C1"] = 'Amt Sold'
write_sheet["D1"] = 'Total'

row_counter = 2

for row in read_ws.iter_rows(min_row=2):
    name = row[0].value
    cost = row[1].value
    amt_sold = row[2].value
    total = row[3].value

    write_sheet.cell(row_counter,1).value = name
    write_sheet.cell(row_counter,2).value = cost
    write_sheet.cell(row_counter,3).value = amt_sold
    write_sheet.cell(row_counter,4).value = total

    row_counter += 1

summary_row = row_counter + 1

write_sheet['B' + str(summary_row)] = 'Total'
write_sheet['B' + str(summary_row)].font = Font(size=16, bold=True)

write_sheet['C' + str(summary_row)] = '=SUM(C2:C' + str(row_counter) + ')'
write_sheet['D' + str(summary_row)] = '=SUM(D2:D' + str(row_counter) + ')'

summary_row2 = row_counter + 2

write_sheet['C' + str(summary_row2)] = '=Average(C2:C' + str(row_counter) + ')'
write_sheet['D' + str(summary_row2)] = '=Average(D2:D' + str(row_counter) + ')'

write_sheet.column_dimensions['A'].width = 16
write_sheet.column_dimensions['B'].width = 16
write_sheet.column_dimensions['C'].width = 16
write_sheet.column_dimensions['D'].width = 16

for cell in write_sheet['C:C']:
    cell.number_format = "#,##0"

for cell in write_sheet['D:D']:
    cell.number_format = u'"$ "#,##0.00'

wb.save('PythontoExcel.xlsx')